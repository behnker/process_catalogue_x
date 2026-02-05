"""Issue export endpoint — CSV/XLSX download."""

import csv
import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.core.auth import CurrentUser, get_current_user
from src.core.tenancy import get_tenant_db
from src.models.issue_log import IssueLog
from src.schemas.issue_log import IssueExportRequest

router = APIRouter()


@router.post("/export")
async def export_issues(
    body: IssueExportRequest,
    user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Export issues to CSV or XLSX format."""
    query = select(IssueLog).where(
        IssueLog.organization_id == user.organization_id
    )

    # Apply filters
    if body.status_filter:
        query = query.where(IssueLog.issue_status.in_(body.status_filter))
    if body.classification_filter:
        query = query.where(IssueLog.issue_classification.in_(body.classification_filter))
    if body.criticality_filter:
        query = query.where(IssueLog.issue_criticality.in_(body.criticality_filter))
    if body.process_ids:
        query = query.where(IssueLog.process_id.in_(body.process_ids))
    if body.date_from:
        query = query.where(IssueLog.date_raised >= body.date_from)
    if body.date_to:
        query = query.where(IssueLog.date_raised <= body.date_to)

    query = query.order_by(IssueLog.issue_number)
    result = await db.execute(query)
    issues = result.scalars().all()

    if body.format == "xlsx":
        return _export_xlsx(issues)

    return _export_csv(issues)


def _export_csv(issues: list[IssueLog]) -> StreamingResponse:
    """Generate CSV export."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        "ID", "Title", "Classification", "Criticality", "Complexity",
        "Status", "Process Ref", "Process Name", "Date Raised",
        "Target Date", "Resolution Date", "Resolution Summary",
        "Opportunity Flag", "Opportunity Status",
    ])

    # Data rows
    for issue in issues:
        writer.writerow([
            issue.display_id,
            issue.title,
            issue.issue_classification,
            issue.issue_criticality,
            issue.issue_complexity,
            issue.issue_status,
            issue.process_ref,
            issue.process_name,
            issue.date_raised.isoformat() if issue.date_raised else "",
            issue.target_resolution_date.isoformat() if issue.target_resolution_date else "",
            issue.actual_resolution_date.isoformat() if issue.actual_resolution_date else "",
            issue.resolution_summary or "",
            "Yes" if issue.opportunity_flag else "No",
            issue.opportunity_status or "",
        ])

    output.seek(0)
    today = date.today().isoformat()

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=ops_issues_{today}.csv"
        },
    )


def _export_xlsx(issues: list[IssueLog]) -> StreamingResponse:
    """Generate XLSX export (requires openpyxl)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="XLSX export requires openpyxl. Install with: pip install openpyxl",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Issue Log"

    # Header row with bold font
    headers = [
        "ID", "Title", "Classification", "Criticality", "Complexity",
        "Status", "Process Ref", "Process Name", "Date Raised",
        "Target Date", "Resolution Date", "Resolution Summary",
        "Opportunity Flag", "Opportunity Status",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Data rows
    for row_num, issue in enumerate(issues, 2):
        ws.cell(row=row_num, column=1, value=issue.display_id)
        ws.cell(row=row_num, column=2, value=issue.title)
        ws.cell(row=row_num, column=3, value=issue.issue_classification)
        ws.cell(row=row_num, column=4, value=issue.issue_criticality)
        ws.cell(row=row_num, column=5, value=issue.issue_complexity)
        ws.cell(row=row_num, column=6, value=issue.issue_status)
        ws.cell(row=row_num, column=7, value=issue.process_ref)
        ws.cell(row=row_num, column=8, value=issue.process_name)
        ws.cell(row=row_num, column=9, value=issue.date_raised)
        ws.cell(row=row_num, column=10, value=issue.target_resolution_date)
        ws.cell(row=row_num, column=11, value=issue.actual_resolution_date)
        ws.cell(row=row_num, column=12, value=issue.resolution_summary or "")
        ws.cell(row=row_num, column=13, value="Yes" if issue.opportunity_flag else "No")
        ws.cell(row=row_num, column=14, value=issue.opportunity_status or "")

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    today = date.today().isoformat()

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=ops_issues_{today}.xlsx"
        },
    )
